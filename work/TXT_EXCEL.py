import openpyxl


def writetoxlsx():
    '''
    将优惠小区8508模块的基站代码从TXT文件存放到EXCEL表格里
    :return:
    '''

    data = open('E:\\WORK\\优惠小区\\6ka0000151.txt', 'r')
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet

    # print(data)
    i = 1  # 注意：'cell'函数中行列起始值为1
    for line in data:
        print(line[0:5])
        print(line[6:15])
        outws.cell(column=2, row=i, value="%s" % line[0:5])
        outws.cell(column=3, row=i, value="%s" % line[6:15])
        i += 1
    savexlsx = "E:\\WORK\\优惠小区\\6ka0000151.xlsx"
    outwb.save(savexlsx)  # 保存结果
    data.close()
#
writetoxlsx()
